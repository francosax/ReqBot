from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def write_excel_file(df, excel_file):
    # Carica il workbook e accedi al foglio di lavoro da aggiornare
    book = load_workbook(excel_file)
    writer = book['MACHINE COMP. MATRIX']

    # font_high = Font(color=Color(rgb='00FF0000')) # Rosso
    # font_medium = Font(color=Color(rgb='00FFFF00')) # Giallo
    # font_low = Font(color=Color(rgb='0000FF00'))

    fill_high = PatternFill(start_color=Color(rgb='00FF0000'), end_color=Color(rgb='00FF0000'), fill_type='solid')
    fill_medium = PatternFill(start_color=Color(rgb='00FFFF00'), end_color=Color(rgb='00FFFF00'), fill_type='solid')
    fill_low = PatternFill(start_color=Color(rgb='0000FF00'), end_color=Color(rgb='0000FF00'), fill_type='solid')

    # Update the Compliance Matrix with the requirements
    for i, (value1, value2, value3, value4, value5, priority) in enumerate(
            zip(df.index, df['Page'], df['Label Number'], df['Description'], df['Priority'], df['Priority']), start=5):

        writer['A{}'.format(i)] = value1
        writer['B{}'.format(i)] = value2
        writer['C{}'.format(i)] = value3
        writer['D{}'.format(i)] = value4
        writer['H{}'.format(i)] = value5

        if priority == 'high':
            writer['H{}'.format(i)].fill = fill_high
        elif priority == 'medium':
            writer['H{}'.format(i)].fill = fill_medium
        elif priority == 'low':
            writer['H{}'.format(i)].fill = fill_low

    # Aggiungi qui le definizioni della Data Validation
    dv1 = DataValidation(type="list",
                         formula1='"Technical,Procedure,Legal,SW,HW,Safety,Documentation,Safety,Warning,N.A."',
                         allow_blank=True)
    dv1.add('I5:I1048576')

    dv2 = DataValidation(type="list", formula1='"Machine,Product,Company"', allow_blank=True)
    dv2.add('J5:J1048576')

    dv3 = DataValidation(type="list", formula1='"Concept,UTM,UTS,UTE,SW,Testing,Process,Assembly,Logistic,'
                                               'Quality,PM,Purchasing,Sales,Service"', allow_blank=True)
    dv3.add('K5:K1048576')

    dv4 = DataValidation(type="list", formula1='"Approved,Rejected,In discussion,Acquired"', allow_blank=True)
    dv4.add('M5:M1048576')  # SOTTO CONVERTIRE J CON M

    dv5 = DataValidation(type="list", formula1='"yes,partially,no"', allow_blank=True)
    dv5.add('N5:N1048576')  # SOTTO SOSTITURE K con N

    dv6 = DataValidation(type="list", formula1='"easy,medium,hard"', allow_blank=True)
    dv6.add('O5:O1048576')  # SOTTO SOSTITUIRE CON O

    dv7 = DataValidation(type="list", formula1='"completed,on going,blocked,failed"', allow_blank=True)
    dv7.add('U5:U1048576')

    dv8 = DataValidation(type="list", formula1='"compliant,not compliant,partially compliant"', allow_blank=True)
    dv8.add('W5:W1048576')

    writer.add_data_validation(dv1)
    writer.add_data_validation(dv2)
    writer.add_data_validation(dv3)
    writer.add_data_validation(dv4)
    writer.add_data_validation(dv5)
    writer.add_data_validation(dv6)
    writer.add_data_validation(dv7)
    writer.add_data_validation(dv8)

    for i in range(5, writer.max_row + 1):
        col_h = get_column_letter(8)  # H
        col_m = get_column_letter(13)  # M
        col_n = get_column_letter(14)  # N
        col_o = get_column_letter(15)  # O
        col_p = get_column_letter(16)  # P

        formula = f'=ROUND((((IF({col_h}{i}="high", 3, IF({col_h}{i}="medium", 2, IF({col_h}{i}="low", 1, 0))) * 4/3 +'\
                  f'IF({col_n}{i}="yes", 1, IF({col_n}{i}="partially", 2, IF({col_n}{i}="no", 3, 0))) * 3/3 +' \
                  f'IF({col_o}{i}="hard", 3, IF({col_o}{i}="medium", 2, IF({col_o}{i}="easy", 1, 0))) * 2/3) -3)*' \
                  f'IF({col_m}{i}="Approved", 1, IF({col_m}{i}="Rejected", 0, IF({col_m}{i}="In discussion", 1,\
                  IF({col_m}{i}= "Acquired", 1, 0))))), 2)'

        # Assegna la formula alla cella nella colonna M
        writer[f'{col_p}{i}'].value = None  # Rimuovi il valore esistente, se presente
        writer[f'{col_p}{i}'].value = formula  # Inserisci la formula
        writer[f'{col_p}{i}'].number_format = '0'

    # Salva il workbook
    book.save(excel_file)
