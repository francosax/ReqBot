import os
import pytest
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
import tempfile  # <--- ADD THIS IMPORT

# Assuming write_excel_file is in a file named excel_writer.py
# Adjust this import path if your file is named differently or in a subfolder
from excel_writer import write_excel_file


@pytest.fixture
def empty_compliance_matrix_template():
    """
    Creates a temporary, empty Excel template file with the required sheet
    'MACHINE COMP. MATRIX' for testing.
    """
    fd, path = tempfile.mkstemp(suffix=".xlsx")  # This line now has tempfile imported
    os.close(fd)  # Close the file descriptor immediately

    book = Workbook()  # Create a new workbook
    # Remove the default 'Sheet' and add your specific sheet
    if 'Sheet' in book.sheetnames:
        del book['Sheet']
    matrix_sheet = book.create_sheet('MACHINE COMP. MATRIX', 0)  # Create as the first sheet

    # Add dummy headers that align with expected columns for formulas and data writing
    matrix_sheet['A1'] = "REQ_ID"
    matrix_sheet['B1'] = "Page"
    matrix_sheet['C1'] = "Label Number"
    matrix_sheet['D1'] = "Description"
    matrix_sheet['E1'] = "Confidence"  # NEW: Confidence header (v2.1)
    matrix_sheet['J1'] = "Category"  # NEW: Category header (v2.2)
    matrix_sheet['I1'] = "Priority"  # SHIFTED from H to I
    matrix_sheet['N1'] = "Status"  # SHIFTED from M to N
    matrix_sheet['O1'] = "Completeness"  # SHIFTED from N to O
    matrix_sheet['P1'] = "Difficulty"  # SHIFTED from O to P
    matrix_sheet['Q1'] = "Calculated Value"  # SHIFTED from P to Q

    book.save(path)
    yield path  # Provide the path to the test
    os.remove(path)  # Clean up the temporary file after the test


def test_write_excel_file_data_and_priority_fills(empty_compliance_matrix_template):
    """
    Tests if data from DataFrame is written correctly and priority-based fills are applied.
    """
    excel_file = empty_compliance_matrix_template

    data = {
        'Page': [1, 2, 3],
        'Label Number': ['L001', 'L002', 'L003'],
        'Description': ['Req 1', 'Req 2', 'Req 3'],
        'Priority': ['high', 'medium', 'low'],
        'Confidence': [0.95, 0.72, 0.55],  # NEW: Confidence scores
        'Category': ['Functional', 'Safety', 'Performance']  # NEW: Category (v2.2)
    }
    df = pd.DataFrame(data, index=['REQ-A', 'REQ-B', 'REQ-C'])

    write_excel_file(df, excel_file)

    # Load the updated workbook to assert changes
    book = load_workbook(excel_file)
    writer = book['MACHINE COMP. MATRIX']

    # Assert data written
    assert writer['A5'].value == 'REQ-A'
    assert writer['B5'].value == 1
    assert writer['I5'].value == 'high'  # SHIFTED from H to I

    assert writer['A6'].value == 'REQ-B'
    assert writer['B6'].value == 2
    assert writer['I6'].value == 'medium'  # SHIFTED from H to I

    assert writer['A7'].value == 'REQ-C'
    assert writer['B7'].value == 3
    assert writer['I7'].value == 'low'  # SHIFTED from H to I

    # Assert fill colors
    # Hexadecimal values for colors without alpha channel (AARRGGBB -> RRGGBB)
    fill_high_rgb = Color(rgb='FF0000')  # Red
    fill_medium_rgb = Color(rgb='FFFF00')  # Yellow
    fill_low_rgb = Color(rgb='00FF00')  # Green

    # Compare fill objects' start_color.rgb
    assert writer['I5'].fill.start_color.rgb == fill_high_rgb.rgb  # SHIFTED from H to I
    assert writer['I6'].fill.start_color.rgb == fill_medium_rgb.rgb  # SHIFTED from H to I
    assert writer['I7'].fill.start_color.rgb == fill_low_rgb.rgb  # SHIFTED from H to I

    book.close()


def test_write_excel_file_data_validations(empty_compliance_matrix_template):
    """
    Tests if data validations are added to the correct cells.
    Note: Checking the exact rules is complex without parsing XML.
    We'll check if validations exist for the specified ranges.
    """
    excel_file = empty_compliance_matrix_template

    data = {  # Minimal data just to get the function to run
        'Page': [1], 'Label Number': ['L001'], 'Description': ['Desc'], 'Priority': ['high'], 'Confidence': [0.85], 'Category': ['Functional']
    }
    df = pd.DataFrame(data, index=['REQ-X'])

    write_excel_file(df, excel_file)

    book = load_workbook(excel_file)
    writer = book['MACHINE COMP. MATRIX']

    # Assert that data validations exist
    # Check for presence of validation rules on the sheet
    validations = writer.data_validations.dataValidation
    assert len(validations) >= 8  # Expecting 8 data validation rules as defined

    book.close()


def test_write_excel_file_formulas(empty_compliance_matrix_template):
    """
    Tests if formulas are written to the correct cells.
    """
    excel_file = empty_compliance_matrix_template

    data = {
        'Page': [1, 2],
        'Label Number': ['L001', 'L002'],
        'Description': ['Req 1', 'Req 2'],
        'Priority': ['high', 'low'],
        'Confidence': [0.92, 0.68],  # NEW: Confidence scores
        'Category': ['Functional', 'Safety']  # NEW: Category (v2.2)
    }
    df = pd.DataFrame(data, index=['REQ-A', 'REQ-B'])

    write_excel_file(df, excel_file)

    book = load_workbook(excel_file)
    writer = book['MACHINE COMP. MATRIX']

    # Expected column 'Q' (17th column) starting from row 5 - SHIFTED from P (16)
    col_q = get_column_letter(17)

    # Check that a formula string is present
    assert writer[f'{col_q}5'].value is not None
    assert writer[f'{col_q}5'].value.startswith('=ROUND(((')
    assert writer[f'{col_q}6'].value is not None
    assert writer[f'{col_q}6'].value.startswith('=ROUND(((')

    # Check number format
    assert writer[f'{col_q}5'].number_format == '0'

    book.close()


def test_write_excel_file_confidence_display_and_formatting(empty_compliance_matrix_template):
    """
    Tests if confidence scores are written correctly with proper conditional formatting.
    Tests color coding: Green (≥0.8), Yellow (0.6-0.8), Red (<0.6)
    """
    excel_file = empty_compliance_matrix_template

    data = {
        'Page': [1, 2, 3],
        'Label Number': ['L001', 'L002', 'L003'],
        'Description': ['High conf req', 'Medium conf req', 'Low conf req'],
        'Priority': ['high', 'medium', 'low'],
        'Confidence': [0.95, 0.72, 0.55],  # High, Medium, Low confidence
        'Category': ['Functional', 'Safety', 'Performance']  # NEW: Category (v2.2)
    }
    df = pd.DataFrame(data, index=['REQ-A', 'REQ-B', 'REQ-C'])

    write_excel_file(df, excel_file)

    # Load the updated workbook to assert changes
    book = load_workbook(excel_file)
    writer = book['MACHINE COMP. MATRIX']

    # Assert confidence values are written
    assert writer['E5'].value == 0.95
    assert writer['E6'].value == 0.72
    assert writer['E7'].value == 0.55

    # Assert number format is set to decimal
    assert writer['E5'].number_format == '0.00'
    assert writer['E6'].number_format == '0.00'
    assert writer['E7'].number_format == '0.00'

    # Assert fill colors for confidence
    # High confidence (≥0.8): Green
    fill_conf_high_rgb = Color(rgb='00FF00')  # Green
    assert writer['E5'].fill.start_color.rgb == fill_conf_high_rgb.rgb

    # Medium confidence (0.6-0.8): Yellow
    fill_conf_medium_rgb = Color(rgb='FFFF00')  # Yellow
    assert writer['E6'].fill.start_color.rgb == fill_conf_medium_rgb.rgb

    # Low confidence (<0.6): Red
    fill_conf_low_rgb = Color(rgb='FF0000')  # Red
    assert writer['E7'].fill.start_color.rgb == fill_conf_low_rgb.rgb

    book.close()


def test_write_excel_file_auto_filter(empty_compliance_matrix_template):
    """
    Tests if auto-filter is applied to allow filtering requirements.
    """
    excel_file = empty_compliance_matrix_template

    data = {
        'Page': [1, 2, 3],
        'Label Number': ['L001', 'L002', 'L003'],
        'Description': ['Req 1', 'Req 2', 'Req 3'],
        'Priority': ['high', 'medium', 'low'],
        'Confidence': [0.95, 0.72, 0.55],
        'Category': ['Functional', 'Safety', 'Performance']  # NEW: Category (v2.2)
    }
    df = pd.DataFrame(data, index=['REQ-A', 'REQ-B', 'REQ-C'])

    write_excel_file(df, excel_file)

    book = load_workbook(excel_file)
    writer = book['MACHINE COMP. MATRIX']

    # Assert that auto-filter is applied
    assert writer.auto_filter is not None
    assert writer.auto_filter.ref is not None

    # Verify auto-filter starts at row 4 (header row)
    assert 'A4' in writer.auto_filter.ref

    # Verify auto-filter includes data rows
    assert '7' in writer.auto_filter.ref  # Should include row 7 (last data row)

    book.close()