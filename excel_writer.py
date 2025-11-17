import os
from openpyxl import load_workbook, Workbook # Import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd # You'll need pandas to create a DataFrame easily

def write_excel_file(df, excel_file):
    """
    Loads an Excel workbook, updates a specific sheet with data from a DataFrame,
    applies conditional formatting based on priority and confidence, adds data validations,
    inserts formulas, and enables auto-filtering.

    Args:
        df (pandas.DataFrame): The DataFrame containing data to write.
                               Expected columns: 'Page', 'Label Number', 'Description', 'Priority', 'Confidence'.
                               The DataFrame index is used for 'A' column.
        excel_file (str): The path to the existing Excel file to be updated.
    """
    try:
        # Carica il workbook e accedi al foglio di lavoro da aggiornare
        book = load_workbook(excel_file)
        # Ensure the sheet exists, otherwise this will raise a KeyError
        if 'MACHINE COMP. MATRIX' not in book.sheetnames:
            print(f"Error: Sheet 'MACHINE COMP. MATRIX' not found in {excel_file}")
            return

        writer = book['MACHINE COMP. MATRIX']

        # Priority color fills (for column H)
        fill_high = PatternFill(start_color=Color(rgb='00FF0000'), end_color=Color(rgb='00FF0000'), fill_type='solid') # Red
        fill_medium = PatternFill(start_color=Color(rgb='00FFFF00'), end_color=Color(rgb='00FFFF00'), fill_type='solid') # Yellow
        fill_low = PatternFill(start_color=Color(rgb='0000FF00'), end_color=Color(rgb='0000FF00'), fill_type='solid') # Green

        # Confidence color fills (for column E)
        # High confidence: Green (≥0.8), Medium: Yellow (0.6-0.8), Low: Red (<0.6)
        fill_conf_high = PatternFill(start_color=Color(rgb='0000FF00'), end_color=Color(rgb='0000FF00'), fill_type='solid') # Green
        fill_conf_medium = PatternFill(start_color=Color(rgb='00FFFF00'), end_color=Color(rgb='00FFFF00'), fill_type='solid') # Yellow
        fill_conf_low = PatternFill(start_color=Color(rgb='00FF0000'), end_color=Color(rgb='00FF0000'), fill_type='solid') # Red

        # Update the Compliance Matrix with the requirements
        # Starting from row 5
        for i, (value1, value2, value3, value4, value5, value6) in enumerate(
                zip(df.index, df['Page'], df['Label Number'], df['Description'], df['Priority'], df['Confidence']), start=5):

            writer[f'A{i}'] = value1         # df.index
            writer[f'B{i}'] = value2         # Page
            writer[f'C{i}'] = value3         # Label Number
            writer[f'D{i}'] = value4         # Description
            writer[f'E{i}'] = value6         # Confidence (NEW)
            writer[f'I{i}'] = value5         # Priority (SHIFTED from H to I due to Confidence)

            # Apply fill based on priority (column I - shifted from H)
            priority = str(value5).lower() # Ensure priority is string and lowercase for comparison
            if priority == 'high':
                writer[f'I{i}'].fill = fill_high
            elif priority == 'medium':
                writer[f'I{i}'].fill = fill_medium
            elif priority == 'low':
                writer[f'I{i}'].fill = fill_low

            # Apply fill based on confidence (column E)
            # Format confidence as percentage and apply color coding
            try:
                confidence_value = float(value6)
                writer[f'E{i}'].number_format = '0.00'  # Format as decimal

                if confidence_value >= 0.8:
                    writer[f'E{i}'].fill = fill_conf_high  # Green for high confidence
                elif confidence_value >= 0.6:
                    writer[f'E{i}'].fill = fill_conf_medium  # Yellow for medium confidence
                else:
                    writer[f'E{i}'].fill = fill_conf_low  # Red for low confidence
            except (ValueError, TypeError):
                # If confidence is not a valid number, skip formatting
                pass

        # Aggiungi qui le definizioni della Data Validation
        # NOTE: All columns shifted right by 1 due to Confidence column in E
        # Max row for Excel is 1048576, using this for full column validation
        dv1 = DataValidation(type="list",
                             formula1='"Technical,Procedure,Legal,SW,HW,Safety,Documentation,Safety,Warning,N.A."',
                             allow_blank=True)
        dv1.add('J5:J1048576')  # Shifted from I to J

        dv2 = DataValidation(type="list", formula1='"Machine,Product,Company"', allow_blank=True)
        dv2.add('K5:K1048576')  # Shifted from J to K

        dv3 = DataValidation(type="list", formula1='"Concept,UTM,UTS,UTE,SW,Testing,Process,Assembly,Logistic,Quality,PM,Purchasing,Sales,Service"', allow_blank=True)
        dv3.add('L5:L1048576')  # Shifted from K to L

        dv4 = DataValidation(type="list", formula1='"Approved,Rejected,In discussion,Acquired"', allow_blank=True)
        dv4.add('N5:N1048576')  # Shifted from M to N

        dv5 = DataValidation(type="list", formula1='"yes,partially,no"', allow_blank=True)
        dv5.add('O5:O1048576')  # Shifted from N to O

        dv6 = DataValidation(type="list", formula1='"easy,medium,hard"', allow_blank=True)
        dv6.add('P5:P1048576')  # Shifted from O to P

        dv7 = DataValidation(type="list", formula1='"completed,on going,blocked,failed"', allow_blank=True)
        dv7.add('V5:V1048576')  # Shifted from U to V

        dv8 = DataValidation(type="list", formula1='"compliant,not compliant,partially compliant"', allow_blank=True)
        dv8.add('X5:X1048576')  # Shifted from W to X

        writer.add_data_validation(dv1)
        writer.add_data_validation(dv2)
        writer.add_data_validation(dv3)
        writer.add_data_validation(dv4)
        writer.add_data_validation(dv5)
        writer.add_data_validation(dv6)
        writer.add_data_validation(dv7)
        writer.add_data_validation(dv8)

        # Apply formulas
        # NOTE: All columns shifted right by 1 due to Confidence column in E
        for i in range(5, writer.max_row + 1):
            col_i = get_column_letter(9)  # I (Priority - shifted from H)
            col_n = get_column_letter(14) # N (Status - shifted from M)
            col_o = get_column_letter(15) # O (Completeness - shifted from N)
            col_p = get_column_letter(16) # P (Difficulty - shifted from O)
            col_q = get_column_letter(17) # Q (Output column for formula - shifted from P)

            # Construct the formula string. Be careful with double quotes inside
            # the formula if they are part of string literals for Excel.
            # Using single quotes for Python string and escaping double quotes for Excel string literals.
            formula = (
                f'=ROUND((('
                f'(IF({col_i}{i}="high", 3, IF({col_i}{i}="medium", 2, IF({col_i}{i}="low", 1, 0))) * 4/3) + ' # Priority
                f'(IF({col_o}{i}="yes", 1, IF({col_o}{i}="partially", 2, IF({col_o}{i}="no", 3, 0))) * 3/3) + ' # Completeness
                f'(IF({col_p}{i}="hard", 3, IF({col_p}{i}="medium", 2, IF({col_p}{i}="easy", 1, 0))) * 2/3)' # Difficulty
                f') - 3) * '
                f'IF({col_n}{i}="Approved", 1, IF({col_n}{i}="Rejected", 0, IF({col_n}{i}="In discussion", 1, IF({col_n}{i}="Acquired", 1, 0))))'
                f'), 2)'
            )

            writer[f'{col_q}{i}'].value = None  # Clear existing value
            writer[f'{col_q}{i}'].value = formula  # Insert the formula
            writer[f'{col_q}{i}'].number_format = '0' # Ensure it's formatted as a number

        # Add auto-filter to allow filtering by confidence and other columns
        # Apply filter to row 4 (header row) - assumes headers are in row 4
        if writer.max_row >= 4:
            # Auto-filter from column A to the last used column
            last_col = get_column_letter(writer.max_column)
            writer.auto_filter.ref = f'A4:{last_col}{writer.max_row}'
            print(f"Auto-filter applied to range A4:{last_col}{writer.max_row}")

        # Salva il workbook
        book.save(excel_file)
        print(f"Excel file '{excel_file}' updated successfully.")

    except FileNotFoundError:
        print(f"Error: The file '{excel_file}' was not found.")
    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == '__main__':
    excel_template_path = "template_compliance_matrix.xlsx"

    print(f"Script is running from: {os.getcwd()}")
    print(f"Attempting to create/update file at: {os.path.abspath(excel_template_path)}")

    if not os.path.exists(excel_template_path):
        print(f"Creating a dummy template: {excel_template_path}")
        dummy_book = Workbook() # <--- CORRECTED: Use openpyxl.Workbook() for new workbook
        dummy_book.create_sheet('MACHINE COMP. MATRIX', 0) # Create at the first position
        # Remove the default 'Sheet' created by Workbook() constructor
        if 'Sheet' in dummy_book.sheetnames:
            del dummy_book['Sheet']

        dummy_sheet = dummy_book['MACHINE COMP. MATRIX']
        # Add some headers that match your expected data structure
        dummy_sheet['A1'] = "REQ_ID"
        dummy_sheet['B1'] = "Page"
        dummy_sheet['C1'] = "Label Number"
        dummy_sheet['D1'] = "Description"
        dummy_sheet['E1'] = "Confidence"  # NEW: Confidence header
        dummy_sheet['H1'] = "Priority"
        dummy_sheet['M1'] = "Status"
        dummy_sheet['N1'] = "Completeness"
        dummy_sheet['O1'] = "Difficulty"
        dummy_sheet['P1'] = "Calculated Value" # For the formula output

        # Add a few rows of dummy data starting from row 5
        dummy_sheet['A5'] = "DUMMY-001"
        dummy_sheet['B5'] = 1
        dummy_sheet['C5'] = "L001"
        dummy_sheet['D5'] = "Sample Req Desc 1"
        dummy_sheet['E5'] = 0.95  # NEW: High confidence sample
        dummy_sheet['H5'] = "high"
        dummy_sheet['M5'] = "Approved"
        dummy_sheet['N5'] = "yes"
        dummy_sheet['O5'] = "easy"

        dummy_sheet['A6'] = "DUMMY-002"
        dummy_sheet['B6'] = 2
        dummy_sheet['C6'] = "L002"
        dummy_sheet['D6'] = "Sample Req Desc 2"
        dummy_sheet['E6'] = 0.55  # NEW: Low confidence sample
        dummy_sheet['H6'] = "medium"
        dummy_sheet['M6'] = "Rejected"
        dummy_sheet['N6'] = "partially"
        dummy_sheet['O6'] = "hard"

        # Corrected placement for saving the dummy book
        dummy_book.save(excel_template_path)
        print("Dummy template created with initial data and headers.")
    else:
        print(f"File '{excel_template_path}' already exists. Updating it.")


    data = {
        'Page': [1, 1, 2, 3],
        'Label Number': ['L001', 'L002', 'L003', 'L004'],
        'Description': ['Login feature', 'Registration feature', 'Data logging', 'Error handling'],
        'Priority': ['high', 'medium', 'low', 'high'],
        'Confidence': [0.95, 0.72, 0.55, 0.88]  # NEW: Confidence scores for testing
    }
    df = pd.DataFrame(data, index=['REQ-001', 'REQ-002', 'REQ-003', 'REQ-004'])

    print("Attempting to write to Excel file...")
    write_excel_file(df, excel_template_path)
    print("Script finished.")

    if os.path.exists(excel_template_path):
        print(f"File successfully saved at: {os.path.abspath(excel_template_path)}")
    else:
        print("File was not found after script execution, check for errors above.")